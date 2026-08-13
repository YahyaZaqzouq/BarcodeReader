# -*- coding: utf-8 -*-
"""توليد صفحة نظام الباركود من ملف الإكسيل
يقرأ data.xlsx (أو أي ملف .xlsx في نفس المجلد) ويكتب index.html
"""
import os, re, glob, json, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))

def find_source():
    for name in ['data.xlsx', 'بيانات.xlsx']:
        p = os.path.join(HERE, name)
        if os.path.exists(p):
            return p
    files = [f for f in glob.glob(os.path.join(HERE, '*.xlsx'))
             if os.path.basename(f) not in ('template.xlsx',)]
    if files:
        return files[0]
    raise SystemExit('ERROR: no .xlsx file found in folder')

def main():
    src = find_source()
    print('reading:', src)
    wb = load_workbook(src, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    seen_codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None and row[2] is None:
            continue
        code = str(row[0]).strip() if row[0] is not None else ''
        bar  = str(row[1]).strip() if row[1] is not None else ''
        name = str(row[2]).strip() if row[2] is not None else ''
        if bar in ('0', '', 'None', 'nan'):
            bar = ''
        if not code and not name and not bar:
            continue
        rows.append([bar, code, name])
    data_literal = json.dumps(rows, ensure_ascii=False, separators=(',', ':'))
    tpl = open(os.path.join(HERE, 'template.html'), encoding='utf-8').read()
    out_html = tpl.replace('__DATA__', data_literal)
    out_path = os.path.join(HERE, 'index.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(out_html)
    print('OK: index.html written | items:', len(rows))

if __name__ == '__main__':
    main()
